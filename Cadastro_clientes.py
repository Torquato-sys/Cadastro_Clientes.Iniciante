import openpyxl
from openpyxl import Workbook
import pathlib


arquivo = pathlib.Path("Clientes.xlsx")

if arquivo.exists():
    pass
else:
    arquivo=Workbook()
    folha = arquivo.active
    folha['A1'] = "Nome do cliente"
    folha['B1'] = "Idade"
    folha['C1'] = "Genero"
    folha['D1'] = "Contato"
    folha['E1'] = "Endereço"

    arquivo.save('Clientes.xlsx')

    nome = input('Nome completo: ')
    idade = int(input('Idade: '))
    genero = input('Genero:')
    contato = int(input('Contato: '))
    endereco = input('Endereco: ')
try:
    arquivo = openpyxl.load_workbook('Clientes.xlsx')
    folha = arquivo.active
    folha.cell(column=1, row=folha.max_row+1, value=nome)
    folha.cell(column=2, row=folha.max_row, value=idade)
    folha.cell(column=3, row=folha.max_row, value=genero)
    folha.cell(column=4, row=folha.max_row, value=contato)
    folha.cell(column=5, row=folha.max_row, value=endereco)

    arquivo.save(r'Clientes.xlsx')

    print('Dados salvo com sucesso!')
except:
    print('Algo deu errado tente novamente.')
